#!/usr/bin/env python3
"""
polymath_pipeline.py
====================

Consolidated evaluation + table-construction pipeline for the multilingual
mathematical-reasoning study (PolyMath-style, difficulty-weighted accuracy).

This single module bundles every data/metric/table-building step used in the
project. It deliberately contains **no plotting code** -- only the
computations that turn raw model-output spreadsheets into score tables,
combined comparison tables, extra-metric tables (format compliance, output
language, reasoning/answer lengths, optional perplexity) and LaTeX tables.

--------------------------------------------------------------------------------
INPUT FORMAT
--------------------------------------------------------------------------------
A *results folder* contains one spreadsheet per language: ``<lang>.xlsx``
(e.g. ``en.xlsx``, ``de.xlsx``, ``hi.xlsx`` ...). Each spreadsheet has one
worksheet per difficulty level -- ``low``, ``medium``, ``high``, ``top`` --
with columns::

    id | answer | questions_translated |
        <model>__raw_answer | <model>__internal_reasoning | ...

* ``answer`` (or ``answer_translated``) is the gold solution.
* For every model there is a ``<model>__raw_answer`` column and, optionally, a
  ``<model>__internal_reasoning`` column. Some models instead embed their
  chain-of-thought inside ``raw_answer`` between ``<think>...</think>`` tags;
  this is reconciled automatically.

A *prompt-strategy root* (for :func:`combine_scores`) contains several such
results folders, one per prompting condition, e.g.::

    root/
      base/         en.xlsx de.xlsx ...
      base_encot/   en.xlsx de.xlsx ...
      bt/           en.xlsx ...

--------------------------------------------------------------------------------
METRICS
--------------------------------------------------------------------------------
* **Per-level accuracy** -- exact match of the last ``\\boxed{...}`` in a
  model's (reasoning + answer) text against the gold answer, with light
  LaTeX-aware normalisation and a numeric fallback.
* **Difficulty-weighted accuracy (DW-ACC)** -- PolyMath weighting
  ``{low:1, medium:2, high:4, top:8}``; the denominator uses only the levels
  actually present, so subsets are handled correctly.
* **Extra metrics** (per model x level): boxed-format compliance %, dominant
  reasoning/answer language + share (needs ``langdetect``), reasoning/answer
  length in tokens and words (mean ± std), and optional perplexity
  (needs ``transformers``+``torch`` or Apple ``mlx-lm``).

--------------------------------------------------------------------------------
COMMAND-LINE USAGE
--------------------------------------------------------------------------------
    # 1) Score one results folder -> per-language xlsx + JSON summary
    python polymath_pipeline.py score   RESULTS_DIR   -o scores_out/

    # 2) Combine several prompt conditions into one comparison workbook
    python polymath_pipeline.py combine  STRATEGY_ROOT -o combined.xlsx

    # 3) Extra metrics (compliance / language / lengths [/ perplexity])
    python polymath_pipeline.py extra    RESULTS_DIR   -o extra_out/ \\
        [--perplexity --lm-model Qwen/Qwen3-4B --backend auto]

    # 4) Length tables (per-language + grouped LaTeX, length_stats.xlsx)
    python polymath_pipeline.py lengths  RESULTS_DIR   -o lengths_out/

Each subcommand is also a plain importable function (see ``__all__``).

Dependencies: ``openpyxl`` (required); ``langdetect`` (optional, language
columns); ``transformers``+``torch`` or ``mlx-lm`` (optional, perplexity).
"""

from __future__ import annotations

import argparse
import json
import math
import re
import statistics
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import openpyxl

__all__ = [
    "LEVEL_WEIGHTS", "ALL_LEVELS",
    "extract_last_boxed", "is_match",
    "boxed_exact_match_accuracy", "difficulty_weighted_accuracy",
    "evaluate_file", "evaluate_folder", "save_results_xlsx",
    "combine_scores",
    "split_reasoning_answer", "compute_extra_metrics_folder",
    "save_extra_metrics_xlsx",
    "compute_length_stats", "save_length_stats_xlsx",
    "latex_length_table",
]

# ============================================================================
# Constants.
# ============================================================================
LEVEL_WEIGHTS: Dict[str, int] = {"low": 1, "medium": 2, "high": 4, "top": 8}
ALL_LEVELS: Tuple[str, ...] = ("low", "medium", "high", "top")


# ============================================================================
# 1. Boxed-answer extraction + matching.
# ============================================================================
def extract_last_boxed(text: Optional[str]) -> Optional[str]:
    r"""Return the content of the LAST ``\boxed{...}`` block, or ``None``.

    Brace-balanced, so nested braces such as ``\boxed{\frac{1}{2}}`` work.
    The *last* occurrence is used because models often restate the final
    answer in a boxed expression at the end.
    """
    if not text:
        return None
    needle = r"\boxed{"
    last = None
    i = 0
    while True:
        idx = text.find(needle, i)
        if idx == -1:
            break
        start = idx + len(needle)
        depth, j = 1, start
        while j < len(text) and depth > 0:
            ch = text[j]
            if ch == "{":
                depth += 1
            elif ch == "}":
                depth -= 1
            j += 1
        if depth == 0:
            last = text[start:j - 1]
        i = j
    return last


_LATEX_NOISE = (r"\!", r"\,", r"\;", r"\:", r"\ ",
                r"\left", r"\right", r"\displaystyle", r"\textstyle")


def _normalise(s: Optional[str]) -> Optional[str]:
    if s is None:
        return None
    s = str(s).strip()
    s = re.sub(r"^\$+|\$+$", "", s).strip()      # strip surrounding $...$
    for tok in _LATEX_NOISE:
        s = s.replace(tok, "")
    s = re.sub(r"\s+", "", s)
    while len(s) >= 2 and s.startswith("{") and s.endswith("}"):
        s = s[1:-1]
    return s


def _try_numeric(s: str) -> Optional[float]:
    cleaned = (s.replace(",", "").replace("{", "").replace("}", "")
               .replace("\\%", "").replace("%", ""))
    try:
        return float(cleaned)
    except (ValueError, TypeError):
        return None


def is_match(predicted: Optional[str], gold: Optional[str]) -> bool:
    """Exact match with light LaTeX normalisation + numeric fallback."""
    p, g = _normalise(predicted), _normalise(gold)
    if p is None or g is None:
        return False
    if p == g:
        return True
    pn, gn = _try_numeric(p), _try_numeric(g)
    if pn is not None and gn is not None:
        return abs(pn - gn) < 1e-9
    return False


# ============================================================================
# Shared sheet helpers.
# ============================================================================
def _find_gold_index(headers) -> int:
    for name in ("answer", "answer_translated"):
        if name in headers:
            return headers.index(name)
    raise ValueError(f"No answer column found in headers: {headers}")


def _padded(row: tuple, n: int) -> tuple:
    """openpyxl trims trailing empties; pad rows back to header width."""
    return row + (None,) * (n - len(row)) if len(row) < n else row


def _model_raw_reasoning_columns(headers) -> Dict[str, Dict[str, int]]:
    """model -> {'raw': idx, 'reasoning': idx} for the columns present."""
    out: Dict[str, Dict[str, int]] = {}
    for i, h in enumerate(headers):
        if not h:
            continue
        if h.endswith("__raw_answer"):
            out.setdefault(h[:-len("__raw_answer")], {})["raw"] = i
        elif h.endswith("__internal_reasoning"):
            out.setdefault(h[:-len("__internal_reasoning")], {})["reasoning"] = i
    return out


# ============================================================================
# 2. Per-level accuracy + difficulty-weighted accuracy.
# ============================================================================
def boxed_exact_match_accuracy(sheet) -> Dict[str, float]:
    r"""``correct / n`` per model for one worksheet.

    The prediction is the last ``\boxed{...}`` in the concatenation of the
    model's ``raw_answer`` and ``internal_reasoning`` cells.
    """
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {}
    headers, body = rows[0], rows[1:]
    n = len(headers)
    answer_idx = _find_gold_index(headers)

    model_cols: Dict[str, List[int]] = {}
    for i, h in enumerate(headers):
        if h and ("__raw_answer" in h or "__internal_reasoning" in h):
            model_cols.setdefault(h.split("__")[0], []).append(i)

    accs: Dict[str, float] = {}
    for model, cols in model_cols.items():
        correct = total = 0
        for raw_row in body:
            row = _padded(raw_row, n)
            gold = row[answer_idx]
            if gold is None:
                continue
            combined = "\n".join(str(row[c]) for c in cols if row[c])
            correct += int(is_match(extract_last_boxed(combined), gold))
            total += 1
        accs[model] = correct / total if total else 0.0
    return accs


def difficulty_weighted_accuracy(
    per_level_acc: Dict[str, Dict[str, float]],
    levels: Optional[Iterable[str]] = None,
) -> Dict[str, float]:
    """Aggregate ``{level:{model:acc}}`` with PolyMath difficulty weights.

    The denominator sums only the weights of levels that are present for a
    model, so evaluating on a subset of levels stays correctly normalised.
    """
    levels = list(levels) if levels is not None else list(per_level_acc.keys())
    models = sorted({m for lvl in levels if lvl in per_level_acc
                     for m in per_level_acc[lvl]})
    out: Dict[str, float] = {}
    for m in models:
        num, denom = 0.0, 0
        for lvl in levels:
            if lvl in per_level_acc and m in per_level_acc[lvl]:
                w = LEVEL_WEIGHTS[lvl]
                num += w * per_level_acc[lvl][m]
                denom += w
        out[m] = num / denom if denom else 0.0
    return out


def evaluate_file(path: str | Path) -> Dict[str, Dict[str, float]]:
    """Return ``{level: {model: accuracy}}`` for one xlsx file."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: Dict[str, Dict[str, float]] = {}
    for sn in wb.sheetnames:
        if sn in LEVEL_WEIGHTS:
            out[sn] = boxed_exact_match_accuracy(wb[sn])
    wb.close()
    return out


def evaluate_folder(
    folder: str | Path,
    levels: Optional[Iterable[str]] = None,
) -> Dict[str, Dict]:
    """Evaluate every ``<lang>.xlsx`` in ``folder``.

    Returns ``{lang: {"per_level": {...}, "aggregate": {model: DW-ACC}}}``.
    """
    folder = Path(folder)
    out: Dict[str, Dict] = {}
    for xlsx in sorted(folder.glob("*.xlsx")):
        if xlsx.name.startswith("~$"):     # skip Excel lock files
            continue
        per_lvl = evaluate_file(xlsx)
        out[xlsx.stem] = {
            "per_level": per_lvl,
            "aggregate": difficulty_weighted_accuracy(per_lvl, levels=levels),
        }
    return out


def save_results_xlsx(
    results: Dict[str, Dict],
    out_dir: str | Path,
    as_percent: bool = True,
) -> List[Path]:
    """Write one ``<lang>.xlsx`` per language.

    Per-level sheets: ``model | accuracy_pct``. Plus an ``aggregate`` sheet:
    ``model | DW_ACC_accuracy_pct | mean_accuracy_pct``.
    """
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    scale = 100.0 if as_percent else 1.0
    fmt = "0.0" if as_percent else "0.0000"
    label = "accuracy_pct" if as_percent else "accuracy"
    written: List[Path] = []

    for lang, data in results.items():
        per_level, aggregate = data["per_level"], data["aggregate"]
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for lvl in ALL_LEVELS:
            if lvl not in per_level:
                continue
            ws = wb.create_sheet(title=lvl)
            ws.append(["model", label])
            for model in sorted(per_level[lvl]):
                ws.append([model, per_level[lvl][model] * scale])
                ws.cell(row=ws.max_row, column=2).number_format = fmt
        ws = wb.create_sheet(title="aggregate")
        ws.append(["model", f"DW_ACC_{label}", f"mean_{label}"])
        models = sorted({m for d in per_level.values() for m in d})
        for model in models:
            scores = [per_level[lvl][model] for lvl in per_level
                      if model in per_level[lvl]]
            mean_acc = sum(scores) / len(scores) if scores else 0.0
            ws.append([model, aggregate.get(model, 0.0) * scale, mean_acc * scale])
            ws.cell(row=ws.max_row, column=2).number_format = fmt
            ws.cell(row=ws.max_row, column=3).number_format = fmt
        path = out_dir / f"{lang}.xlsx"
        wb.save(path)
        written.append(path)
    return written


# ============================================================================
# 3. Combine several prompt conditions into one comparison workbook.
# ============================================================================
def _read_level_sheet(path: Path, sheet_name: str) -> Dict[str, float]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close(); return {}
    rows = list(wb[sheet_name].iter_rows(values_only=True)); wb.close()
    if not rows:
        return {}
    h = list(rows[0])
    if "model" not in h:
        return {}
    mi = h.index("model")
    vi = 1 if len(h) > 1 else None
    out = {}
    for r in rows[1:]:
        if r and r[mi] is not None and vi is not None and \
                isinstance(r[vi], (int, float)):
            out[r[mi]] = r[vi]
    return out


def _read_aggregate_sheet(path: Path, value_col: str) -> Dict[str, float]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "aggregate" not in wb.sheetnames:
        wb.close(); return {}
    rows = list(wb["aggregate"].iter_rows(values_only=True)); wb.close()
    if not rows:
        return {}
    h = list(rows[0])
    if "model" not in h or value_col not in h:
        return {}
    mi, vi = h.index("model"), h.index(value_col)
    return {r[mi]: r[vi] for r in rows[1:]
            if r and r[mi] is not None and isinstance(r[vi], (int, float))}


def _discover_conditions(root: Path) -> List[str]:
    return sorted(d.name for d in root.iterdir()
                  if d.is_dir() and any(d.glob("*.xlsx")))


def combine_scores(
    root: str | Path,
    output: str | Path = "combined.xlsx",
    conditions: Optional[Iterable[str]] = None,
    lang_first: Optional[Iterable[str]] = ("en",),
    value_col: str = "DW_ACC_accuracy_pct",
) -> Path:
    """Build one workbook comparing prompt conditions side by side.

    Produces 5 sheets (``low``/``medium``/``high``/``top``/``aggregate``).
    Rows are models; columns are grouped by language, with one sub-column per
    condition (two header rows, merged language headers, frozen panes).

    ``root`` holds one sub-folder per condition (auto-discovered if
    ``conditions`` is ``None``). ``lang_first`` pins languages (e.g. English)
    to the leftmost column group.
    """
    root = Path(root)
    conditions = list(conditions) if conditions else _discover_conditions(root)
    if not conditions:
        raise ValueError(f"No condition sub-folders with xlsx found in {root}")

    # data[sheet][lang][condition] -> {model: value}
    data: Dict[str, Dict[str, Dict[str, Dict[str, float]]]] = {}
    sheets = list(ALL_LEVELS) + ["aggregate"]
    for sheet in sheets:
        data[sheet] = {}
    for cond in conditions:
        for xlsx in sorted((root / cond).glob("*.xlsx")):
            if xlsx.name.startswith("~$"):
                continue
            lang = xlsx.stem
            for sheet in ALL_LEVELS:
                d = _read_level_sheet(xlsx, sheet)
                if d:
                    data[sheet].setdefault(lang, {})[cond] = d
            agg = _read_aggregate_sheet(xlsx, value_col)
            if agg:
                data["aggregate"].setdefault(lang, {})[cond] = agg

    lang_first_list = list(lang_first) if lang_first else None

    def col_order(sheet_data: Dict) -> List[str]:
        langs = list(sheet_data.keys())
        if lang_first_list:
            head = [L for L in lang_first_list if L in langs]
            tail = sorted(L for L in langs if L not in head)
            return head + tail
        return sorted(langs)

    def all_models(sheet_data: Dict) -> List[str]:
        ms = set()
        for lang in sheet_data:
            for cond in sheet_data[lang]:
                ms.update(sheet_data[lang][cond])
        return sorted(ms)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet in sheets:
        sd = data[sheet]
        ws = wb.create_sheet(title=sheet)
        langs = col_order(sd)
        models = all_models(sd)
        # Header row 1: language spanning len(conditions) columns.
        # Header row 2: condition names.
        ws.cell(row=1, column=1, value="model")
        ws.cell(row=2, column=1, value="")
        col = 2
        for lang in langs:
            start = col
            for cond in conditions:
                ws.cell(row=2, column=col, value=cond)
                col += 1
            end = col - 1
            c = ws.cell(row=1, column=start, value=lang)
            if end > start:
                ws.merge_cells(start_row=1, start_column=start,
                               end_row=1, end_column=end)
            c.alignment = openpyxl.styles.Alignment(horizontal="center")
        # Body.
        for ri, model in enumerate(models, start=3):
            ws.cell(row=ri, column=1, value=model)
            col = 2
            for lang in langs:
                for cond in conditions:
                    v = sd.get(lang, {}).get(cond, {}).get(model)
                    cell = ws.cell(row=ri, column=col, value=v)
                    if isinstance(v, (int, float)):
                        cell.number_format = "0.0"
                    col += 1
        ws.freeze_panes = "B3"
    output = Path(output)
    wb.save(output)
    return output


# ============================================================================
# 4. Reasoning / answer splitting + extra metrics.
# ============================================================================
_THINK_RE = re.compile(r"<think>(.*?)</think>", re.DOTALL | re.IGNORECASE)
_THINK_OPEN_RE = re.compile(r"<think>(.*)$", re.DOTALL | re.IGNORECASE)


def split_reasoning_answer(
    raw_answer: Optional[str], reasoning: Optional[str],
) -> Tuple[Optional[str], Optional[str]]:
    """Return ``(reasoning_text, answer_text)`` reconciling two output styles.

    Either a dedicated ``internal_reasoning`` column is present, or the chain
    of thought is embedded in ``raw_answer`` inside ``<think>...</think>``
    (possibly unclosed if the generation was truncated).
    """
    if reasoning:
        ans = raw_answer
        if raw_answer:
            ans = _THINK_RE.sub("", str(raw_answer))
            ans = _THINK_OPEN_RE.sub("", ans).strip() or None
        return str(reasoning), ans
    if raw_answer:
        s = str(raw_answer)
        m = _THINK_RE.search(s)
        if m:
            return (m.group(1).strip() or None,
                    _THINK_RE.sub("", s).strip() or None)
        m2 = _THINK_OPEN_RE.search(s)
        if m2:
            return m2.group(1).strip() or None, None
        return None, s
    return None, None


def _word_count(text: Optional[str]) -> Optional[int]:
    if not text:
        return None
    return len(str(text).split())


def _mode_and_share(labels: List[Optional[str]]) -> Tuple[Optional[str], float]:
    vals = [x for x in labels if x]
    if not vals:
        return None, 0.0
    counts: Dict[str, int] = {}
    for v in vals:
        counts[v] = counts.get(v, 0) + 1
    best = max(counts, key=counts.get)
    return best, counts[best] / len(vals)


def _fmt_mean_std(values: List[Optional[float]]) -> str:
    nums = [v for v in values if v is not None and not math.isinf(v)]
    if not nums:
        return ""
    mean = statistics.fmean(nums)
    std = statistics.pstdev(nums) if len(nums) > 1 else 0.0
    return f"{mean:.1f} ± {std:.1f}"


class _LangDetector:
    """Thin lazy wrapper around ``langdetect`` (optional dependency)."""

    def __init__(self) -> None:
        self.available = False
        try:
            from langdetect import detect, DetectorFactory  # type: ignore
            DetectorFactory.seed = 0
            self._detect = detect
            self.available = True
        except Exception:
            self._detect = None

    def detect(self, text: Optional[str]) -> Optional[str]:
        if not self.available or not text:
            return None
        try:
            return self._detect(str(text)[:1000])
        except Exception:
            return None


def compute_extra_metrics_folder(
    folder: str | Path,
    use_perplexity: bool = False,
    lm_model: str = "Qwen/Qwen3-4B",
    backend: str = "auto",
    verbose: bool = True,
) -> Dict[str, Dict]:
    """Compute extra metrics for every ``<lang>.xlsx`` in ``folder``.

    Returns ``{lang: {level: {model: {metric: value}}}}`` with metrics:
    ``boxed_compliance_pct``, ``reasoning_lang`` (+``_pct``),
    ``answer_lang`` (+``_pct``), ``reasoning_len_words``/``answer_len_words``
    as ``"mean ± std"`` strings, and (if ``use_perplexity``) token lengths and
    perplexity. Perplexity needs ``transformers``+``torch`` or ``mlx-lm``;
    when unavailable those cells read ``"NA (no LM)"``.

    NOTE: the heavy perplexity scorer is intentionally lazy/optional so this
    module imports and runs with only ``openpyxl`` installed.
    """
    folder = Path(folder)
    langdet = _LangDetector()
    if use_perplexity and verbose and not langdet.available:
        print("[extra] langdetect unavailable -> language cols 'NA (no langdetect)'")

    scorer = None
    if use_perplexity:
        scorer = _try_load_perplexity(lm_model, backend, verbose)

    out: Dict[str, Dict] = {}
    for xlsx in sorted(folder.glob("*.xlsx")):
        if xlsx.name.startswith("~$"):
            continue
        lang = xlsx.stem
        wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
        out[lang] = {}
        for sn in wb.sheetnames:
            if sn not in LEVEL_WEIGHTS:
                continue
            out[lang][sn] = _compute_sheet_extra(wb[sn], langdet, scorer)
        wb.close()
    return out


def _compute_sheet_extra(sheet, langdet, scorer) -> Dict[str, Dict]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {}
    headers, body = rows[0], rows[1:]
    n = len(headers)
    mcols = _model_raw_reasoning_columns(headers)

    result: Dict[str, Dict] = {}
    for model, cols in mcols.items():
        raw_i = cols.get("raw")
        rea_i = cols.get("reasoning")
        n_inst = 0
        boxed_ok = 0
        rea_langs: List[Optional[str]] = []
        ans_langs: List[Optional[str]] = []
        rea_words: List[Optional[float]] = []
        ans_words: List[Optional[float]] = []
        rea_tok: List[Optional[float]] = []
        ans_tok: List[Optional[float]] = []
        rea_ppl: List[Optional[float]] = []
        ans_ppl: List[Optional[float]] = []

        for raw_row in body:
            row = _padded(raw_row, n)
            raw = row[raw_i] if raw_i is not None else None
            rea_col = row[rea_i] if rea_i is not None else None
            if raw is None and rea_col is None:
                continue
            n_inst += 1
            reasoning, answer = split_reasoning_answer(raw, rea_col)

            combined = "\n".join(x for x in (str(raw) if raw else "",
                                             str(rea_col) if rea_col else "")
                                 if x)
            boxed_ok += int(extract_last_boxed(combined) is not None)

            rea_langs.append(langdet.detect(reasoning))
            ans_langs.append(langdet.detect(answer))
            rea_words.append(_word_count(reasoning))
            ans_words.append(_word_count(answer))

            if scorer is not None:
                rea_tok.append(scorer.count_tokens(reasoning))
                ans_tok.append(scorer.count_tokens(answer))
                rea_ppl.append(scorer.perplexity(reasoning))
                ans_ppl.append(scorer.perplexity(answer))

        rlang, rshare = _mode_and_share(rea_langs)
        alang, ashare = _mode_and_share(ans_langs)
        tok_sentinel = "NA (no LM)" if scorer is None else None

        result[model] = {
            "n_instances": n_inst,
            "boxed_compliance_pct": 100.0 * boxed_ok / n_inst if n_inst else 0.0,
            "reasoning_lang": rlang or ("NA (no langdetect)"
                                        if not langdet.available else None),
            "reasoning_lang_pct": round(100.0 * rshare, 1),
            "answer_lang": alang or ("NA (no langdetect)"
                                     if not langdet.available else None),
            "answer_lang_pct": round(100.0 * ashare, 1),
            "reasoning_len_words": _fmt_mean_std(rea_words),
            "answer_len_words": _fmt_mean_std(ans_words),
            "reasoning_len_tokens": tok_sentinel or _fmt_mean_std(rea_tok),
            "answer_len_tokens": tok_sentinel or _fmt_mean_std(ans_tok),
            "reasoning_perplexity": tok_sentinel or _fmt_mean_std(rea_ppl),
            "answer_perplexity": tok_sentinel or _fmt_mean_std(ans_ppl),
        }
    return result


_EXTRA_COLUMNS = [
    ("n_instances", "n_instances"),
    ("boxed_compliance_pct", "boxed_compliance_%"),
    ("reasoning_lang", "reasoning_lang"),
    ("reasoning_lang_pct", "reasoning_lang_%"),
    ("answer_lang", "answer_lang"),
    ("answer_lang_pct", "answer_lang_%"),
    ("reasoning_len_tokens", "reasoning_len_tokens (mean±std)"),
    ("reasoning_len_words", "reasoning_len_words (mean±std)"),
    ("reasoning_perplexity", "reasoning_perplexity (mean±std)"),
    ("answer_len_tokens", "answer_len_tokens (mean±std)"),
    ("answer_len_words", "answer_len_words (mean±std)"),
    ("answer_perplexity", "answer_perplexity (mean±std)"),
]


def save_extra_metrics_xlsx(
    metrics: Dict[str, Dict], out_dir: str | Path,
) -> List[Path]:
    """Write one ``<lang>.xlsx`` per language with one sheet per level."""
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    written: List[Path] = []
    headers = ["model"] + [disp for _k, disp in _EXTRA_COLUMNS]
    for lang, per_level in metrics.items():
        wb = openpyxl.Workbook(); wb.remove(wb.active)
        for lvl in ALL_LEVELS:
            if lvl not in per_level:
                continue
            ws = wb.create_sheet(title=lvl)
            ws.append(headers)
            for model in sorted(per_level[lvl]):
                d = per_level[lvl][model]
                ws.append([model] + [d.get(k) for k, _disp in _EXTRA_COLUMNS])
        path = out_dir / f"{lang}.xlsx"
        wb.save(path)
        written.append(path)
    return written


def _try_load_perplexity(lm_model: str, backend: str, verbose: bool):
    """Return a loaded perplexity scorer, or ``None`` if deps are missing."""
    try:
        scorer = _PerplexityScorer(lm_model, backend=backend, verbose=verbose)
        return scorer if scorer.load() else None
    except Exception as exc:  # noqa: BLE001
        if verbose:
            print(f"[extra] perplexity disabled: {exc}")
        return None


def _looks_like_mlx(name: str) -> bool:
    n = name.lower()
    return (n.startswith("mlx-community/") or "/mlx-" in n
            or n.endswith("-mlx") or (n.endswith("bit") and "mlx" in n))


class _PerplexityScorer:
    """Optional causal-LM perplexity scorer (HF Transformers or Apple MLX).

    Loaded lazily; if the relevant libraries/model are unavailable, ``load()``
    returns ``False`` and the caller falls back to ``"NA (no LM)"`` sentinels.
    """

    def __init__(self, model_name: str, max_chars: int = 4000,
                 backend: str = "auto", verbose: bool = False) -> None:
        self.model_name = model_name
        self.max_chars = max_chars
        self.verbose = verbose
        self.backend = ("mlx" if _looks_like_mlx(model_name) else "hf") \
            if backend == "auto" else backend
        self._ok = False
        self._tok = self._model = self._torch = self._mx = None

    def load(self) -> bool:
        if self._ok:
            return True
        try:
            if self.backend == "mlx":
                import mlx.core as mx           # type: ignore
                from mlx_lm import load          # type: ignore
                self._mx = mx
                self._model, self._tok = load(self.model_name)
            else:
                import torch                      # type: ignore
                from transformers import (AutoTokenizer,  # type: ignore
                                          AutoModelForCausalLM)
                self._torch = torch
                self._tok = AutoTokenizer.from_pretrained(self.model_name)
                dtype = torch.float16 if torch.cuda.is_available() else None
                self._model = AutoModelForCausalLM.from_pretrained(
                    self.model_name, torch_dtype=dtype)
                self._model.eval()
                if torch.cuda.is_available():
                    self._model.to("cuda")
            self._ok = True
        except Exception as exc:  # noqa: BLE001
            if self.verbose:
                print(f"[extra] LM '{self.model_name}' "
                      f"(backend={self.backend}) failed: "
                      f"{type(exc).__name__}: {exc}")
            self._ok = False
        return self._ok

    def count_tokens(self, text: Optional[str]) -> Optional[int]:
        if not self._ok or not text:
            return None
        if self.backend == "mlx":
            return len(self._tok.encode(str(text)))
        return len(self._tok(str(text), add_special_tokens=False)["input_ids"])

    def perplexity(self, text: Optional[str]) -> Optional[float]:
        if not self._ok or not text:
            return None
        s = str(text)[: self.max_chars]
        if not s.strip():
            return None
        if self.backend == "mlx":
            import mlx.nn as nn                   # type: ignore
            ids = self._tok.encode(s)[:2048]
            if len(ids) < 2:
                return None
            tok = self._mx.array(ids)[None]
            logits = self._model(tok[:, :-1])
            loss = nn.losses.cross_entropy(
                logits.reshape(-1, logits.shape[-1]), tok[:, 1:].reshape(-1))
            try:
                return math.exp(float(self._mx.mean(loss)))
            except OverflowError:
                return float("inf")
        torch = self._torch
        enc = self._tok(s, return_tensors="pt", truncation=True, max_length=2048)
        ids = enc["input_ids"]
        if ids.shape[1] < 2:
            return None
        ids = ids.to(next(self._model.parameters()).device)
        with torch.no_grad():
            loss = float(self._model(ids, labels=ids).loss)
        try:
            return math.exp(loss)
        except OverflowError:
            return float("inf")


# ============================================================================
# 5. Length statistics + LaTeX tables.
# ============================================================================
def compute_length_stats(
    folder: str | Path, unit: str = "words",
) -> Dict[str, Dict[str, Dict[str, Dict[str, float]]]]:
    """Per (lang, model, level) mean+std of reasoning/answer/total lengths.

    ``unit="words"`` uses whitespace token counts (no extra dependencies).
    Returns ``{lang: {model: {level: {part: {"mean":.., "std":..}}}}}`` where
    ``part`` in ``{"reasoning","answer","total"}`` and ``level`` also includes
    ``"all"`` (pooled across levels).
    """
    folder = Path(folder)
    out: Dict = {}
    for xlsx in sorted(folder.glob("*.xlsx")):
        if xlsx.name.startswith("~$"):
            continue
        lang = xlsx.stem
        wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
        # collect raw per-sample word counts: [model][level][part] -> list
        raw: Dict[str, Dict[str, Dict[str, List[int]]]] = {}
        for sn in wb.sheetnames:
            if sn not in LEVEL_WEIGHTS:
                continue
            rows = list(wb[sn].iter_rows(values_only=True))
            if not rows:
                continue
            headers = rows[0]
            n = len(headers)
            mcols = _model_raw_reasoning_columns(headers)
            for raw_row in rows[1:]:
                row = _padded(raw_row, n)
                for model, cols in mcols.items():
                    raw_v = row[cols["raw"]] if "raw" in cols else None
                    rea_v = row[cols["reasoning"]] if "reasoning" in cols else None
                    if raw_v is None and rea_v is None:
                        continue
                    reasoning, answer = split_reasoning_answer(raw_v, rea_v)
                    rc = _word_count(reasoning) or 0
                    ac = _word_count(answer) or 0
                    md = raw.setdefault(model, {}).setdefault(
                        sn, {"reasoning": [], "answer": [], "total": []})
                    md["reasoning"].append(rc)
                    md["answer"].append(ac)
                    md["total"].append(rc + ac)
        wb.close()

        out[lang] = {}
        for model, levels in raw.items():
            out[lang][model] = {}
            pooled = {"reasoning": [], "answer": [], "total": []}
            for lvl in ALL_LEVELS:
                if lvl not in levels:
                    continue
                out[lang][model][lvl] = {}
                for part in ("reasoning", "answer", "total"):
                    vals = levels[lvl][part]
                    pooled[part].extend(vals)
                    out[lang][model][lvl][part] = {
                        "mean": statistics.fmean(vals) if vals else 0.0,
                        "std": statistics.pstdev(vals) if len(vals) > 1 else 0.0,
                    }
            out[lang][model]["all"] = {
                part: {
                    "mean": statistics.fmean(pooled[part]) if pooled[part] else 0.0,
                    "std": statistics.pstdev(pooled[part]) if len(pooled[part]) > 1 else 0.0,
                } for part in ("reasoning", "answer", "total")
            }
    return out


def save_length_stats_xlsx(
    stats: Dict, out_path: str | Path, part: str = "total", unit: str = "words",
) -> Path:
    """Write a ``length_stats.xlsx`` (one sheet per language).

    Columns: ``model | {low,medium,high,top,all}_mean | ..._std | unit`` for
    the chosen ``part`` (``reasoning``/``answer``/``total``).
    """
    out_path = Path(out_path)
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    cols = (["model"]
            + [f"{lvl}_mean" for lvl in list(ALL_LEVELS) + ["all"]]
            + [f"{lvl}_std" for lvl in list(ALL_LEVELS) + ["all"]]
            + ["unit"])
    for lang, models in stats.items():
        ws = wb.create_sheet(title=lang[:31])
        ws.append(cols)
        for model in sorted(models):
            md = models[model]
            means = [md.get(lvl, {}).get(part, {}).get("mean") for lvl in
                     list(ALL_LEVELS) + ["all"]]
            stds = [md.get(lvl, {}).get(part, {}).get("std") for lvl in
                    list(ALL_LEVELS) + ["all"]]
            ws.append([model] + means + stds + [unit])
    wb.save(out_path)
    return out_path


def latex_length_table(
    stats: Dict, languages: Sequence[str], part: str = "total",
    caption: str = "Output length (words, mean).",
    label: str = "tab:lengths",
) -> str:
    """Return a booktabs LaTeX table of per-model 'all' mean length, one
    column per requested language."""
    langs = [l for l in languages if l in stats]
    models = sorted({m for l in langs for m in stats[l]})
    col_spec = "l" + "r" * len(langs)
    lines = [r"\begin{table}[t]", r"\centering", r"\small",
             rf"\caption{{{caption}}}", rf"\label{{{label}}}",
             rf"\begin{{tabular}}{{{col_spec}}}", r"\toprule",
             r"\textbf{Model} & " + " & ".join(rf"\textbf{{{l}}}" for l in langs)
             + r" \\", r"\midrule"]
    for m in models:
        cells = []
        for l in langs:
            v = stats[l].get(m, {}).get("all", {}).get(part, {}).get("mean")
            cells.append(f"{v:.0f}" if isinstance(v, (int, float)) else "--")
        safe = m.replace("_", r"\_")
        lines.append(rf"\texttt{{{safe}}} & " + " & ".join(cells) + r" \\")
    lines += [r"\bottomrule", r"\end{tabular}", r"\end{table}", ""]
    return "\n".join(lines)


# ============================================================================
# CLI.
# ============================================================================
def _cli() -> None:
    ap = argparse.ArgumentParser(
        description="PolyMath multilingual evaluation + table pipeline.")
    sub = ap.add_subparsers(dest="cmd", required=True)

    p = sub.add_parser("score", help="Per-level + DW-ACC scores for a folder.")
    p.add_argument("results_dir")
    p.add_argument("-o", "--out-dir", default="scores_out")
    p.add_argument("--fraction", action="store_true",
                   help="Write fractions in [0,1] instead of percentages.")
    p.add_argument("--json", action="store_true",
                   help="Also dump the raw results dict as JSON.")

    p = sub.add_parser("combine", help="Combine prompt conditions -> 1 workbook.")
    p.add_argument("strategy_root")
    p.add_argument("-o", "--output", default="combined.xlsx")
    p.add_argument("--conditions", nargs="*", default=None)
    p.add_argument("--no-en-first", action="store_true")

    p = sub.add_parser("extra", help="Extra metrics (compliance/lang/length).")
    p.add_argument("results_dir")
    p.add_argument("-o", "--out-dir", default="extra_out")
    p.add_argument("--perplexity", action="store_true")
    p.add_argument("--lm-model", default="Qwen/Qwen3-4B")
    p.add_argument("--backend", choices=["auto", "hf", "mlx"], default="auto")

    p = sub.add_parser("lengths", help="Length stats xlsx + LaTeX table.")
    p.add_argument("results_dir")
    p.add_argument("-o", "--out-dir", default="lengths_out")
    p.add_argument("--part", choices=["reasoning", "answer", "total"],
                   default="total")

    args = ap.parse_args()

    if args.cmd == "score":
        results = evaluate_folder(args.results_dir)
        paths = save_results_xlsx(results, args.out_dir,
                                  as_percent=not args.fraction)
        print(f"Wrote {len(paths)} files to {args.out_dir}")
        if args.json:
            jp = Path(args.out_dir) / "results.json"
            jp.write_text(json.dumps(results, indent=2))
            print(f"Wrote {jp}")

    elif args.cmd == "combine":
        out = combine_scores(
            args.strategy_root, args.output,
            conditions=args.conditions,
            lang_first=None if args.no_en_first else ("en",))
        print(f"Wrote {out}")

    elif args.cmd == "extra":
        metrics = compute_extra_metrics_folder(
            args.results_dir, use_perplexity=args.perplexity,
            lm_model=args.lm_model, backend=args.backend)
        paths = save_extra_metrics_xlsx(metrics, args.out_dir)
        print(f"Wrote {len(paths)} files to {args.out_dir}")

    elif args.cmd == "lengths":
        out_dir = Path(args.out_dir); out_dir.mkdir(parents=True, exist_ok=True)
        stats = compute_length_stats(args.results_dir)
        xp = save_length_stats_xlsx(stats, out_dir / "length_stats.xlsx",
                                    part=args.part)
        langs = sorted(stats.keys())
        tex = latex_length_table(stats, langs, part=args.part)
        (out_dir / "length_table.tex").write_text(tex)
        print(f"Wrote {xp} and {out_dir/'length_table.tex'}")


if __name__ == "__main__":
    _cli()
